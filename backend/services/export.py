import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def export_to_excel(data: List[Dict[str, Any]], headers: List[str], title: str = "Reporte") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_num, row_data in enumerate(data, 5):
        for col_num, header in enumerate(headers, 1):
            key = header.lower().replace(' ', '_').replace('ó', 'o').replace('í', 'i').replace('á', 'a').replace('é', 'e').replace('ú', 'u')
            value = row_data.get(key, row_data.get(header, ''))
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
    
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in range(1, len(data) + 5):
            cell = ws.cell(row=row, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def export_to_pdf(data: List[Dict[str, Any]], headers: List[str], title: str = "Reporte") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=1
    )
    elements.append(Paragraph(title, title_style))
    
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        alignment=1
    )
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
    elements.append(Spacer(1, 20))
    
    table_data = [headers]
    for row in data:
        row_values = []
        for header in headers:
            key = header.lower().replace(' ', '_').replace('ó', 'o').replace('í', 'i').replace('á', 'a').replace('é', 'e').replace('ú', 'u')
            value = row.get(key, row.get(header, ''))
            row_values.append(str(value) if value else '')
        table_data.append(row_values)
    
    if len(table_data) > 1:
        col_widths = [doc.width / len(headers)] * len(headers)
        table = Table(table_data, colWidths=col_widths)
        
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ])
        table.setStyle(style)
        elements.append(table)
    else:
        elements.append(Paragraph("No hay datos para mostrar", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def generate_sales_report_data(sales) -> tuple:
    headers = ["Numero", "Fecha", "Total", "Metodo de Pago", "Usuario", "Tienda"]
    data = []
    for sale in sales:
        payment_methods = ", ".join([p.payment_method for p in sale.payments]) if sale.payments else "N/A"
        data.append({
            "numero": sale.sale_number,
            "fecha": sale.created_at.strftime('%Y-%m-%d %H:%M'),
            "total": f"${float(sale.total):,.0f}",
            "metodo_de_pago": payment_methods,
            "usuario": sale.user.full_name or sale.user.username if sale.user else "N/A",
            "tienda": sale.store.name if sale.store else "N/A"
        })
    return headers, data

def generate_inventory_report_data(products) -> tuple:
    headers = ["Producto", "Marca", "Cantidad", "Precio", "Stock Minimo", "Proveedor"]
    data = []
    for product in products:
        data.append({
            "producto": product.name,
            "marca": product.brand or "N/A",
            "cantidad": product.quantity,
            "precio": f"${float(product.sale_price):,.0f}",
            "stock_minimo": product.min_stock,
            "proveedor": product.supplier or "N/A"
        })
    return headers, data

def generate_expenses_report_data(expenses) -> tuple:
    headers = ["Fecha", "Monto", "Metodo", "Descripcion", "Usuario", "Tienda"]
    data = []
    for expense in expenses:
        data.append({
            "fecha": expense.created_at.strftime('%Y-%m-%d %H:%M'),
            "monto": f"${float(expense.amount):,.0f}",
            "metodo": expense.payment_method,
            "descripcion": expense.description[:50] + "..." if len(expense.description) > 50 else expense.description,
            "usuario": "N/A",
            "tienda": "N/A"
        })
    return headers, data

def generate_cash_closing_report(closing, opening, sales, expenses, transfers_in, transfers_out) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=12,
        alignment=1
    )
    elements.append(Paragraph("CIERRE DE CAJA", title_style))
    elements.append(Spacer(1, 20))
    
    info_data = [
        ["Fecha de Apertura:", opening.opening_date.strftime('%Y-%m-%d %H:%M')],
        ["Fecha de Cierre:", closing.closing_date.strftime('%Y-%m-%d %H:%M')],
        ["Base Inicial:", f"${float(opening.initial_balance):,.0f}"],
        ["Total Ventas:", f"${float(closing.total_sales):,.0f}"],
        ["Total Egresos:", f"${float(closing.total_expenses):,.0f}"],
        ["Transferencias Entrantes:", f"${float(closing.total_transfers_in):,.0f}"],
        ["Transferencias Salientes:", f"${float(closing.total_transfers_out):,.0f}"],
        ["Balance Esperado:", f"${float(closing.expected_balance):,.0f}"],
        ["Balance Real:", f"${float(closing.actual_balance):,.0f}"],
        ["Diferencia:", f"${float(closing.difference):,.0f}"],
    ]
    
    info_table = Table(info_data, colWidths=[200, 200])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F5F5')),
    ]))
    elements.append(info_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
